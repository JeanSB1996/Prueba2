import csv
import pandas as pd
import openpyxl
from recalculo_gratificacion import main as recalc_main

RLI_FILES = [
    '3.2.5 78570410-K Renta Liquida Imponoble y determinación 30%.xlsx',
    '3.2.5 96514060-3 Renta Liquida Imponoble y determinación 30%.xlsx'
]


def extract_profit(path: str) -> float:
    """Extract profit from Renta Liquida workbook."""
    df = pd.read_excel(path, sheet_name='Utilidad Liquida', header=None)
    row = df[df.apply(lambda r: r.astype(str).str.contains('Reparto Utilidades').any(), axis=1)].iloc[0]
    profit_30 = row[5] if not pd.isna(row[5]) else row[8]
    return float(profit_30) / 0.30


def gather_profit(paths) -> float:
    return sum(extract_profit(p) for p in paths)


def update_hallazgos(excel_path: str, profit: float, min_wage: float = 460000.0):
    temp_csv = 'hallazgos_temp.csv'
    recalc_main(excel_path, profit, min_wage, output=temp_csv)

    wb = openpyxl.load_workbook(excel_path)
    if 'Hallazgos' in wb.sheetnames:
        del wb['Hallazgos']
    ws = wb.create_sheet('Hallazgos')
    with open(temp_csv, newline='') as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(excel_path)


def main():
    profit = gather_profit(RLI_FILES)
    update_hallazgos('Analisis Gratificacion 2024 - Prosegur V1.xlsx', profit)


if __name__ == '__main__':
    main()
